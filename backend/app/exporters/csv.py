"""
CSV exporter for spreadsheet-friendly data.

Generates CSV files that can be imported into Excel, Google Sheets, etc.
Multiple sheets exported as separate CSV files or ZIP archive.
"""

from typing import Any, Dict
import csv
import io

from app.exporters import BaseExporter


class CSVExporter(BaseExporter):
    """
    Export analysis results to CSV format.

    Supports multiple export modes:
    - Technologies inventory
    - Dependencies list
    - Metrics summary
    - Research gaps list
    - Combined (all in one with sections)
    """

    def __init__(self, project_analysis: Dict[str, Any], export_type: str = "combined"):
        """
        Initialize CSV exporter.

        Args:
            project_analysis: Analysis results dictionary
            export_type: Type of CSV to generate (technologies, dependencies, metrics, gaps, combined)
        """
        super().__init__(project_analysis)
        self.export_type = export_type

    def export(self) -> str:
        """
        Export analysis to CSV format.

        Returns:
            CSV data as string
        """
        if self.export_type == "technologies":
            return self._export_technologies()
        elif self.export_type == "dependencies":
            return self._export_dependencies()
        elif self.export_type == "metrics":
            return self._export_metrics()
        elif self.export_type == "gaps":
            return self._export_research_gaps()
        else:  # combined
            return self._export_combined()

    def _export_technologies(self) -> str:
        """
        Export technologies to CSV.

        Returns:
            CSV string
        """
        output = io.StringIO()
        writer = csv.writer(output)

        # Header
        writer.writerow(
            [
                "Technology",
                "Version",
                "Category",
                "Confidence (%)",
                "Description",
                "Source File",
            ]
        )

        # Data rows
        for tech in self._get_technology_list():
            writer.writerow(
                [
                    tech.get("name", ""),
                    tech.get("version", ""),
                    tech.get("category", ""),
                    tech.get("confidence", ""),
                    tech.get("description", ""),
                    tech.get("source_file", ""),
                ]
            )

        return output.getvalue()

    def _export_dependencies(self) -> str:
        """
        Export dependencies to CSV.

        Returns:
            CSV string
        """
        output = io.StringIO()
        writer = csv.writer(output)

        # Header
        writer.writerow(
            [
                "Package",
                "Current Version",
                "Latest Version",
                "Status",
                "Severity",
                "Package Manager",
                "File",
            ]
        )

        # Data rows
        for dep in self._get_dependency_list():
            is_outdated = dep.get("is_outdated", False) or dep.get("needs_update", False)
            status = "Outdated" if is_outdated else "Current"

            writer.writerow(
                [
                    dep.get("name", ""),
                    dep.get("current_version", ""),
                    dep.get("latest_version", ""),
                    status,
                    dep.get("severity", ""),
                    dep.get("package_manager", ""),
                    dep.get("file", ""),
                ]
            )

        return output.getvalue()

    def _export_metrics(self) -> str:
        """
        Export code metrics to CSV.

        Returns:
            CSV string
        """
        output = io.StringIO()
        writer = csv.writer(output)

        metrics = self._get_metrics()

        # Header
        writer.writerow(["Metric", "Value"])

        # Data rows
        writer.writerow(["Lines of Code", metrics.get("lines_of_code", 0)])
        writer.writerow(["File Count", metrics.get("file_count", 0)])
        writer.writerow(["Complexity Score", metrics.get("complexity", 0)])
        writer.writerow(["Documentation Coverage (%)", metrics.get("documentation_coverage", 0)])
        writer.writerow(["Test Coverage (%)", metrics.get("test_coverage", 0)])

        # Add language breakdown if available
        languages = metrics.get("languages", {})
        if languages:
            writer.writerow([])
            writer.writerow(["Language", "Lines"])
            for lang, lines in languages.items():
                writer.writerow([lang, lines])

        return output.getvalue()

    def _export_research_gaps(self) -> str:
        """
        Export research gaps to CSV.

        Returns:
            CSV string
        """
        output = io.StringIO()
        writer = csv.writer(output)

        # Header
        writer.writerow(
            [
                "Title",
                "Description",
                "Category",
                "Priority",
                "Estimated Effort",
                "Recommendation",
            ]
        )

        # Data rows
        for gap in self._get_research_gaps_list():
            writer.writerow(
                [
                    gap.get("title", ""),
                    gap.get("description", ""),
                    gap.get("category", ""),
                    gap.get("priority", ""),
                    gap.get("estimated_effort", ""),
                    gap.get("recommendation", ""),
                ]
            )

        return output.getvalue()

    def _export_combined(self) -> str:
        """
        Export all data to single CSV with sections.

        Returns:
            CSV string
        """
        output = io.StringIO()
        writer = csv.writer(output)

        # Metadata section
        writer.writerow(["CommandCenter Analysis Report"])
        writer.writerow(["Project", self.project_path])
        writer.writerow(["Analyzed At", self.analyzed_at])
        writer.writerow(["Analysis Version", self.analysis.get("analysis_version", "2.0.0")])
        writer.writerow([])

        # Summary section
        technologies = self._get_technology_list()
        dependencies = self._get_dependency_list()
        research_gaps = self._get_research_gaps_list()
        outdated_count = sum(1 for d in dependencies if d.get("is_outdated", False))

        writer.writerow(["SUMMARY"])
        writer.writerow(["Metric", "Value"])
        writer.writerow(["Technologies Detected", len(technologies)])
        writer.writerow(["Total Dependencies", len(dependencies)])
        writer.writerow(["Outdated Dependencies", outdated_count])
        writer.writerow(["Research Gaps", len(research_gaps)])
        writer.writerow([])

        # Technologies section
        writer.writerow(["DETECTED TECHNOLOGIES"])
        writer.writerow(["Technology", "Version", "Category", "Confidence (%)", "Description"])
        for tech in technologies:
            writer.writerow(
                [
                    tech.get("name", ""),
                    tech.get("version", ""),
                    tech.get("category", ""),
                    tech.get("confidence", ""),
                    tech.get("description", ""),
                ]
            )
        writer.writerow([])

        # Dependencies section
        writer.writerow(["DEPENDENCIES"])
        writer.writerow(["Package", "Current Version", "Latest Version", "Status", "Severity"])
        for dep in dependencies:
            is_outdated = dep.get("is_outdated", False)
            status = "Outdated" if is_outdated else "Current"
            writer.writerow(
                [
                    dep.get("name", ""),
                    dep.get("current_version", ""),
                    dep.get("latest_version", ""),
                    status,
                    dep.get("severity", ""),
                ]
            )
        writer.writerow([])

        # Metrics section
        writer.writerow(["CODE METRICS"])
        metrics = self._get_metrics()
        writer.writerow(["Metric", "Value"])
        writer.writerow(["Lines of Code", metrics.get("lines_of_code", 0)])
        writer.writerow(["File Count", metrics.get("file_count", 0)])
        writer.writerow(["Complexity Score", metrics.get("complexity", 0)])
        writer.writerow([])

        # Research gaps section
        writer.writerow(["RESEARCH GAPS"])
        writer.writerow(["Title", "Description", "Category", "Priority", "Recommendation"])
        for gap in research_gaps:
            writer.writerow(
                [
                    gap.get("title", ""),
                    gap.get("description", ""),
                    gap.get("category", ""),
                    gap.get("priority", ""),
                    gap.get("recommendation", ""),
                ]
            )

        return output.getvalue()


class ExcelExporter(CSVExporter):
    """
    Export analysis results to Excel format (.xlsx).

    Requires openpyxl library.
    Creates multi-sheet workbook with formatting.
    """

    def export(self) -> bytes:
        """
        Export analysis to Excel format.

        Returns:
            Excel file as bytes

        Raises:
            ImportError: If openpyxl is not installed
        """
        try:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError(
                "openpyxl is required for Excel export. Install with: pip install openpyxl"
            )

        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Create sheets
        self._create_summary_sheet(wb)
        self._create_technologies_sheet(wb)
        self._create_dependencies_sheet(wb)
        self._create_metrics_sheet(wb)
        self._create_research_gaps_sheet(wb)

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def _create_summary_sheet(self, wb: Any) -> None:
        """Create summary sheet.

        Args:
            wb: openpyxl Workbook instance
        """
        from openpyxl.styles import Font, PatternFill

        ws = wb.create_sheet("Summary", 0)

        # Title
        ws["A1"] = "CommandCenter Analysis Report"
        ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
        ws["A1"].fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")

        # Metadata
        ws["A3"] = "Project:"
        ws["B3"] = self.project_path
        ws["A4"] = "Analyzed At:"
        ws["B4"] = self.analyzed_at

        # Summary metrics
        technologies = self._get_technology_list()
        dependencies = self._get_dependency_list()
        research_gaps = self._get_research_gaps_list()
        outdated_count = sum(1 for d in dependencies if d.get("is_outdated", False))

        ws["A6"] = "Summary Metrics"
        ws["A6"].font = Font(bold=True)
        ws["A7"] = "Technologies Detected"
        ws["B7"] = len(technologies)
        ws["A8"] = "Total Dependencies"
        ws["B8"] = len(dependencies)
        ws["A9"] = "Outdated Dependencies"
        ws["B9"] = outdated_count
        ws["A10"] = "Research Gaps"
        ws["B10"] = len(research_gaps)

        # Auto-width
        for col in ["A", "B"]:
            ws.column_dimensions[col].width = 25

    def _create_technologies_sheet(self, wb: Any) -> None:
        """Create technologies sheet.

        Args:
            wb: openpyxl Workbook instance
        """
        from openpyxl.styles import Font, PatternFill

        ws = wb.create_sheet("Technologies")

        # Header
        headers = ["Technology", "Version", "Category", "Confidence (%)", "Description"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")

        # Data
        for row, tech in enumerate(self._get_technology_list(), 2):
            ws.cell(row, 1, tech.get("name", ""))
            ws.cell(row, 2, tech.get("version", ""))
            ws.cell(row, 3, tech.get("category", ""))
            ws.cell(row, 4, tech.get("confidence", ""))
            ws.cell(row, 5, tech.get("description", ""))

        # Auto-width
        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 20

    def _create_dependencies_sheet(self, wb: Any) -> None:
        """Create dependencies sheet.

        Args:
            wb: openpyxl Workbook instance
        """
        from openpyxl.styles import Font, PatternFill

        ws = wb.create_sheet("Dependencies")

        # Header
        headers = ["Package", "Current Version", "Latest Version", "Status", "Severity"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")

        # Data
        for row, dep in enumerate(self._get_dependency_list(), 2):
            is_outdated = dep.get("is_outdated", False)
            status = "Outdated" if is_outdated else "Current"

            ws.cell(row, 1, dep.get("name", ""))
            ws.cell(row, 2, dep.get("current_version", ""))
            ws.cell(row, 3, dep.get("latest_version", ""))
            status_cell = ws.cell(row, 4, status)
            ws.cell(row, 5, dep.get("severity", ""))

            # Highlight outdated
            if is_outdated:
                status_cell.fill = PatternFill(
                    start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"
                )

        # Auto-width
        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 18

    def _create_metrics_sheet(self, wb: Any) -> None:
        """Create metrics sheet.

        Args:
            wb: openpyxl Workbook instance
        """
        from openpyxl.styles import Font

        ws = wb.create_sheet("Metrics")

        metrics = self._get_metrics()

        # Header
        ws["A1"] = "Metric"
        ws["B1"] = "Value"
        ws["A1"].font = Font(bold=True)
        ws["B1"].font = Font(bold=True)

        # Data
        row = 2
        ws.cell(row, 1, "Lines of Code")
        ws.cell(row, 2, metrics.get("lines_of_code", 0))
        row += 1
        ws.cell(row, 1, "File Count")
        ws.cell(row, 2, metrics.get("file_count", 0))
        row += 1
        ws.cell(row, 1, "Complexity Score")
        ws.cell(row, 2, metrics.get("complexity", 0))

        # Auto-width
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 15

    def _create_research_gaps_sheet(self, wb: Any) -> None:
        """Create research gaps sheet.

        Args:
            wb: openpyxl Workbook instance
        """
        from openpyxl.styles import Font, PatternFill

        ws = wb.create_sheet("Research Gaps")

        # Header
        headers = ["Title", "Description", "Category", "Priority", "Recommendation"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")

        # Data
        for row, gap in enumerate(self._get_research_gaps_list(), 2):
            ws.cell(row, 1, gap.get("title", ""))
            ws.cell(row, 2, gap.get("description", ""))
            ws.cell(row, 3, gap.get("category", ""))
            ws.cell(row, 4, gap.get("priority", ""))
            ws.cell(row, 5, gap.get("recommendation", ""))

        # Auto-width
        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 25


def export_to_csv(project_analysis: Dict[str, Any], export_type: str = "combined") -> str:
    """
    Convenience function to export analysis to CSV format.

    Args:
        project_analysis: Analysis results dictionary
        export_type: Type of CSV (technologies, dependencies, metrics, gaps, combined)

    Returns:
        CSV data as string
    """
    exporter = CSVExporter(project_analysis, export_type)
    return exporter.export()


def export_to_excel(project_analysis: Dict[str, Any]) -> bytes:
    """
    Convenience function to export analysis to Excel format.

    Args:
        project_analysis: Analysis results dictionary

    Returns:
        Excel file as bytes

    Raises:
        ImportError: If openpyxl is not installed
    """
    exporter = ExcelExporter(project_analysis)
    return exporter.export()
